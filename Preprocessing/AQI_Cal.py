from openpyxl import load_workbook
import xlwings
import pandas as pd
import time
import csv

df = pd.read_csv("pre_delhi_data.csv")
df_temp = df.iloc[:, [3, 4, 6, 7]]
s = time.time()
aqi_list = []

l = [0, 106, 212, 318, 424, 530, 636, 742, 848, 954, 1060]
batch_init = 0
batch_end = 1060

for i in range(batch_init, batch_end):
    wb = load_workbook(filename='AQI_Cal.xlsx')
    ws = wb["Sheet1"]
    l = df_temp.loc[9540+i]
    ws["C8"].value = l[0]
    ws["C10"].value = l[1]
    ws["C12"].value = l[3]
    ws["C20"].value = l[2]
    wb.save("AQI_Cal.xlsx")
    wbxl = xlwings.Book('AQI_Cal.xlsx')
    aqi_val = wbxl.sheets['Sheet1'].range('G11').value
    aqi_list.append([int(aqi_val)])
    wbxl.save()
    wbxl.app.quit()
e = time.time()
print(e-s)

with open("aqi.csv", 'a', newline="") as f:
    writer = csv.writer(f)
    # writer.writerow(["AQI"])
    writer.writerows(aqi_list)
